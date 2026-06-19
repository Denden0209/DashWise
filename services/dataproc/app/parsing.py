"""File readers → array-of-arrays. pandas/openpyxl/pdfplumber do the heavy,
robust reading; the cube/schema ports then operate on the AoA exactly like the
browser does. Handles 500k+ row CSVs comfortably."""
from __future__ import annotations

import io
import json
import re
from typing import Any, Dict, List, Optional, Tuple

MAX_SHEETS = 20


def _coerce(v: Any) -> Any:
    """Convert numpy scalars / NaN to native python so JSON stays clean."""
    if v is None:
        return None
    if hasattr(v, "item"):          # numpy scalar
        try:
            v = v.item()
        except Exception:           # noqa: BLE001
            pass
    if isinstance(v, float) and v != v:   # NaN
        return None
    return v


def _df_to_aoa(df) -> List[List[Any]]:
    return [[_coerce(c) for c in row] for row in df.itertuples(index=False, name=None)]


# ── CSV ────────────────────────────────────────────────────
def read_csv(data: bytes) -> List[List[Any]]:
    import pandas as pd
    text = data.decode("utf-8", errors="replace")
    df = pd.read_csv(io.StringIO(text), sep=None, engine="python",
                     header=None, dtype=object, keep_default_na=True,
                     skip_blank_lines=True, on_bad_lines="skip")
    # dtype=object keeps strings; coerce obvious numerics for the cube/schema.
    return _df_to_aoa(df)


# ── Excel ──────────────────────────────────────────────────
def read_excel(data: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    sheets: List[Dict[str, Any]] = []
    visible_names: List[str] = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":        # skip hidden / veryHidden
            continue
        visible_names.append(ws.title)
        if len(sheets) >= MAX_SHEETS:
            continue
        aoa: List[List[Any]] = [[_coerce(c) for c in row]
                                for row in ws.iter_rows(values_only=True)]
        # Fill merged cells (iter_rows leaves all but the anchor empty).
        for rng in ws.merged_cells.ranges:
            r0, c0, r1, c1 = rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1
            if r0 >= len(aoa) or c0 >= len(aoa[r0] or []):
                continue
            anchor = aoa[r0][c0]
            if anchor in (None, ""):
                continue
            for r in range(r0, min(r1 + 1, len(aoa))):
                for c in range(c0, min(c1 + 1, len(aoa[r]))):
                    if aoa[r][c] in (None, ""):
                        aoa[r][c] = anchor
        if aoa:
            sheets.append({"name": ws.title, "aoa": aoa})
    wb.close()
    return sheets, visible_names


# ── PDF ────────────────────────────────────────────────────
def read_pdf(data: bytes) -> Dict[str, Any]:
    import pdfplumber
    pages: List[str] = []
    best_table: Optional[Dict[str, Any]] = None
    title = ""
    num_pages = 0
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        num_pages = len(pdf.pages)
        try:
            title = str((pdf.metadata or {}).get("Title") or "").strip()
        except Exception:   # noqa: BLE001
            title = ""
        for i, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            if txt.strip():
                pages.append(f"--- Page {i} ---\n{txt.strip()}")
            if best_table is None:
                try:
                    tables = page.extract_tables() or []
                except Exception:   # noqa: BLE001
                    tables = []
                for tb in tables:
                    cleaned = [[(_coerce(c) if c not in (None, "") else "") for c in row] for row in tb]
                    cleaned = [r for r in cleaned if any(str(c).strip() for c in r)]
                    if len(cleaned) >= 4 and len(cleaned[0]) >= 2:
                        headers = [str(h).strip() or f"Column_{j+1}" for j, h in enumerate(cleaned[0])]
                        best_table = {"headers": headers, "rows": cleaned[1:]}
                        break
    body = "\n\n".join(pages).strip()
    meaningful = len(re.sub(r"[^\w]", "", body))
    return {"text": body, "table": best_table, "title": title,
            "num_pages": num_pages, "scanned": meaningful < 20}


# ── JSON ───────────────────────────────────────────────────
_WRAP_KEYS = ("data", "results", "records", "rows", "items")


def _is_record(x: Any) -> bool:
    return isinstance(x, dict)


def extract_records(parsed: Any) -> Optional[List[dict]]:
    if isinstance(parsed, list):
        return [x for x in parsed if _is_record(x)]
    if isinstance(parsed, dict):
        for k in _WRAP_KEYS:
            v = parsed.get(k)
            if isinstance(v, list):
                return [x for x in v if _is_record(x)]
    return None


def read_json(data: bytes) -> Dict[str, Any]:
    text = data.decode("utf-8", errors="replace")
    parsed: Any = None
    records: Optional[List[dict]] = None
    try:
        parsed = json.loads(text)
        records = extract_records(parsed)
    except json.JSONDecodeError:
        objs: List[dict] = []
        ok = True
        for line in (l.strip() for l in text.splitlines() if l.strip()):
            try:
                o = json.loads(line)
                if isinstance(o, dict):
                    objs.append(o)
            except json.JSONDecodeError:
                ok = False
                break
        if ok and objs:
            parsed, records = objs, objs
        else:
            raise ValueError("This file is not valid JSON.")
    return {"parsed": parsed, "records": records}


def records_to_aoa(records: List[dict], cap: int = 5000) -> List[List[Any]]:
    sample = records[:cap]
    headers: List[str] = []
    seen = set()

    def flat(obj: dict) -> dict:
        out: Dict[str, Any] = {}
        for k, v in obj.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    out[f"{k}.{k2}"] = json.dumps(v2) if isinstance(v2, (dict, list)) else v2
            elif isinstance(v, list):
                out[k] = json.dumps(v)
            else:
                out[k] = v
        return out

    flat_records = [flat(r) for r in sample]
    for fr in flat_records:
        for k in fr:
            if k not in seen:
                seen.add(k)
                headers.append(k)
    rows = [[fr.get(h, "") for h in headers] for fr in flat_records]
    return [headers] + rows


# ── TXT delimiter sniff ────────────────────────────────────
def sniff_delimiter(text: str) -> Optional[str]:
    lines = [l for l in text.splitlines() if l.strip()][:20]
    if len(lines) < 3:
        return None
    best_delim, best_cols = None, 1
    for d in ["\t", ",", "|", ";"]:
        counts = [l.count(d) for l in lines]
        mode = max(set(counts), key=counts.count)
        if mode < 1:
            continue
        consistent = sum(1 for c in counts if c == mode)
        if consistent >= len(lines) * 0.8 and mode + 1 > best_cols:
            best_cols, best_delim = mode + 1, d
    return best_delim


def read_txt(data: bytes) -> Dict[str, Any]:
    import pandas as pd
    text = data.decode("utf-8", errors="replace")
    delim = sniff_delimiter(text)
    if delim:
        df = pd.read_csv(io.StringIO(text), sep=re.escape(delim), engine="python",
                         header=None, dtype=object, keep_default_na=True,
                         skip_blank_lines=True, on_bad_lines="skip")
        aoa = _df_to_aoa(df)
        if aoa and len(aoa[0]) >= 2 and len(aoa) >= 4:
            return {"aoa": aoa, "text": text}
    return {"aoa": None, "text": text}
